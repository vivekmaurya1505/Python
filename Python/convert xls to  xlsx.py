'''
import win32com.client as win32
fname = "rC:/Users/ankitaPC/Desktop/vivek_office/Indranil_doc/Sys_Req_Coverage/convert.xls"
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)

wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
wb.Close()                               #FileFormat = 56 is for .xls extension
excel.Application.Quit()
'''

'''
import pyexcel as p
global dest_file_name
p.save_book_as(file_name=r'C:/Users/ankitaPC/Desktop/vivek_office/Indranil_doc/Sys_Req_Coverage/convert.xls',
               n=dest_file_name=r'C:/Users/ankitaPC/Desktop/vivek_office/Indranil_doc/Sys_Req_Coverage/convert.xlsx')
#print(p.save_book_as.dest_file_name)
'''
'''
import xlrd

from openpyxl.workbook import Workbook

from openpyxl.reader.excel import load_workbook, InvalidFileException
filename=r'C:/Users/ankitaPC/Desktop/vivek_office/Indranil_doc/Sys_Req_Coverage/convert.xls'
def open_xls_as_xlsx(filename):

    # first open using xlrd

    book = xlrd.open_workbook(filename)

    index = 0

    nrows, ncols = 0, 0

    while nrows * ncols == 0:

        sheet = book.sheet_by_index(index)

        nrows = sheet.nrows

        ncols = sheet.ncols

        index += 1

    # prepare a xlsx sheet

    book1 = Workbook()

    sheet1 = book1.get_active_sheet()

    for row in xrange(0, nrows):

        for col in xrange(0, ncols):

            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

    return book1

(open_xls_as_xlsx(r'C:/Users/ankitaPC/Desktop/vivek_office/Indranil_doc/Sys_Req_Coverage/convert.xls'))
'''


import pandas as pd
df=pd.read_excel(r'C:/Users/ankitaPC/Desktop/vivek_office/Indranil_doc/Sys_Req_Coverage/convert.xls')
writer=pd.ExcelWriter(r'C:/Users/ankitaPC/Desktop/vivek_office/Indranil_doc/Sys_Req_Coverage/convert.xlsx')
df.to_excel(writer,index=False)
writer.save()

















