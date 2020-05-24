from openpyxl import load_workbook

wb1=load_workbook("file_name1.xlsx")
wb2=load_workbook("file_name2.xlsx")


shee1= wb1.get_sheet_by_name("Work_sheet_name1")
shee2= wb1.get_sheet_by_name("Work_sheet_name2")

for i in range(1,sheet1.max_rows+1):
    for j in range(1,sheet1.max_cols+1):
        sheet2.cell(row=i,column =j).value = sheet1.cell(row = i, column =j).value

wb1.save("file_name1.xlsx")
wb2.save("file_name2.xlsx")
